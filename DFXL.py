import os
import pandas as pd
from openpyxl import Workbook, load_workbook
import numpy as np

class DataFrameXL(pd.DataFrame):
    _metadata = ["_filename", "_sheet_name", "_wb", "_ws", "_styles"]

    @property
    def _constructor(self):
        # pandas usará esto para crear nuevos objetos del mismo tipo
        return DataFrameXL

    def __init__(self, data=None, filename=None, sheet_name="Hoja1", df: pd.DataFrame = None, *args, **kwargs):
        self._filename = filename
        self._sheet_name = sheet_name
        self._styles = {}

        if df is None:
            # Caso 1: inicialización desde Excel
            if filename is not None and isinstance(filename, str) and os.path.exists(filename):
                self._wb = load_workbook(filename)
                if sheet_name in self._wb.sheetnames:
                    self._ws = self._wb[sheet_name]
                else:
                    self._ws = self._wb.create_sheet(sheet_name)

                values = list(self._ws.values)
                if values:
                    columns = values[0]
                    rows = values[1:]
                    df = pd.DataFrame(rows, columns=columns)
                else:
                    df = pd.DataFrame()

                super().__init__(df, *args, **kwargs)
            # Caso 2: inicialización desde datos (cuando pandas llama internamente)
            else:
                if filename is not None and isinstance(filename, str):
                    # Crear workbook nuevo si se pasó filename pero no existe
                    self._wb = Workbook()
                    self._ws = self._wb.active
                    self._ws.title = sheet_name
                else:
                    # Si no hay filename, no inicializamos workbook
                    self._wb = None
                    self._ws = None

                super().__init__(data, *args, **kwargs)
        else:
            if filename is not None:
                # Crear workbook nuevo si se pasó filename
                self._wb = Workbook()
                self._ws = self._wb.active
                self._ws.title = sheet_name
            else:
                # Si no hay filename, no inicializamos workbook
                self._wb = None
                self._ws = None
            super().__init__(df, *args, **kwargs)


    def save(self, filename=None):
        if filename == None:
            filename = self._filename
        # 1. Aplicar estilos antes de guardar
        self.__apply_all_styles()

        # 2. Volcar encabezados en la primera fila
        for j, col_name in enumerate(self.columns):
            self._ws.cell(row=1, column=j+1, value=col_name)

        # 3. Volcar datos fila por fila
        for i in range(len(self)):
            for j, col_name in enumerate(self.columns):
                val = self.iat[i, j]
                self._ws.cell(row=i+2, column=j+1, value=val)

        # 4. Guardar archivo
        if filename:
            self._wb.save(filename)
        else:
            self._wb.save(self._filename)

    def __apply_all_styles(self):
        if not hasattr(self, "_styles"):
            return

        # 1. Estilo global de todo el documento
        if "__document__" in self._styles:
            rules = self._styles["__document__"]
            for row_key, style in rules.items():
                if row_key == "global":
                    # aplicar a todas las celdas (encabezados + datos)
                    for i in range(len(self) + 1):  # +1 para incluir encabezado
                        for j in range(len(self.columns)):
                            cell = self._ws.cell(row=i+1, column=j+1)
                            self._apply_style(cell, style)

        # 2. Estilos por columna/fila como ya tienes
        for col_name, rules in self._styles.items():
            if col_name == "__document__":
                continue
            col_excel = list(self.columns).index(col_name) + 1
            for row_key, style in rules.items():
                if row_key == "global":
                    for i in range(len(self)):
                        cell = self._ws.cell(row=i+2, column=col_excel)
                        self._apply_style(cell, style)
                elif row_key == "header":
                    cell = self._ws.cell(row=1, column=col_excel)
                    self._apply_style(cell, style)
                elif isinstance(row_key, int):
                    cell = self._ws.cell(row=row_key+2, column=col_excel)
                    self._apply_style(cell, style)
                elif isinstance(row_key, slice):
                    for i in range(row_key.start or 0, row_key.stop or len(self)):
                        cell = self._ws.cell(row=i+2, column=col_excel)
                        self._apply_style(cell, style)
                elif isinstance(row_key, (list, np.ndarray)):
                    for i in row_key:
                        cell = self._ws.cell(row=i+2, column=col_excel)
                        self._apply_style(cell, style)


    # Función auxiliar para aplicar estilos
    def _apply_style(self, cell, style):
        if "font" in style:
            cell.font = style["font"]
        if "fill" in style:
            cell.fill = style["fill"]
        if "alignment" in style:
            cell.alignment = style["alignment"]
        if "border" in style:
            cell.border = style["border"]

    @property
    def loc(self):
        base_loc = super().loc
        ws = self._ws
        columns = list(self.columns)

        class _CustomLoc:
            def __getitem__(_, key):
                return base_loc[key]

            def __setitem__(_, key, value):

                # Caso especial: value es dict con data + style
                if isinstance(value, dict) and "data" in value and "style" in value:
                    style = value["style"]
                    data = value["data"]

                    # Asignar datos al DataFrame
                    base_loc[key] = data

                    # Registrar estilos
                    if not hasattr(base_loc.obj, "_styles"):
                        base_loc.obj._styles = {}

                    # Si key es tupla (rows, cols)
                    if isinstance(key, tuple) and len(key) == 2:
                        row_key, col_key = key

                        # Normalizar columnas
                        if isinstance(col_key, str):
                            col_names = [col_key]
                        elif isinstance(col_key, list):
                            col_names = col_key
                        else:
                            col_names = [columns[col_key]]

                        # Normalizar filas
                        if isinstance(row_key, pd.Series) and row_key.dtype == bool:
                            row_indices = row_key[row_key].index.tolist()
                        elif isinstance(row_key, (list, np.ndarray)):
                            row_indices = list(row_key)
                        elif isinstance(row_key, slice):
                            row_indices = list(range(row_key.start or 0, row_key.stop or len(base_loc.obj)))
                        elif isinstance(row_key, int):
                            row_indices = [row_key]
                        else:
                            row_indices = []

                        # Registrar estilos por cada fila afectada
                        for col_name in col_names:
                            if col_name not in base_loc.obj._styles:
                                base_loc.obj._styles[col_name] = {}
                            for idx in row_indices:
                                base_loc.obj._styles[col_name][idx] = style

                    else:
                        # Caso: asignación de columna completa
                        if isinstance(key, str):
                            col_names = [key]
                        elif isinstance(key, list):
                            col_names = key
                        else:
                            col_names = [columns[key]]

                        for col_name in col_names:
                            if col_name not in base_loc.obj._styles:
                                base_loc.obj._styles[col_name] = {}
                            base_loc.obj._styles[col_name]["global"] = style
                else:
                    # Caso normal: solo datos
                    base_loc[key] = value

                # --- Sincronizar Excel ---
                try:
                    for j, col_name in enumerate(columns):
                        for i, val in enumerate(base_loc.obj[col_name]):
                            cell = ws.cell(row=i+2, column=j+1, value=val)

                            # Aplicar estilo si existe
                            if hasattr(base_loc.obj, "_styles") and col_name in base_loc.obj._styles:
                                for row_key, style in base_loc.obj._styles[col_name].items():
                                    if row_key == "global":
                                        base_loc.obj._apply_style(cell, style)
                                    elif isinstance(row_key, int) and row_key == i:
                                        base_loc.obj._apply_style(cell, style)
                                    elif isinstance(row_key, slice) and i in range(row_key.start or 0, row_key.stop or len(base_loc.obj)):
                                        base_loc.obj._apply_style(cell, style)

                except Exception as e:
                    print(f"[ERROR] No se pudo actualizar Excel desde loc: {e}")

            def __getattr__(_, name):
                return getattr(base_loc, name)

        return _CustomLoc()

    @property
    def iloc(self):
        base_iloc = super().iloc
        ws = self._ws
        columns = list(self.columns)

        class _CustomILoc:
            def __getitem__(_, key):
                return base_iloc[key]

            def __setitem__(_, key, value):

                # Caso especial: value es dict con data + style
                if isinstance(value, dict) and "data" in value and "style" in value:
                    style = value["style"]
                    data = value["data"]

                    if isinstance(key, tuple) and len(key) == 2:
                        row_key, col_idx = key
                        col_name = columns[col_idx]

                        # Inicializar estructura de estilos
                        if not hasattr(base_iloc.obj, "_styles"):
                            base_iloc.obj._styles = {}
                        if col_name not in base_iloc.obj._styles:
                            base_iloc.obj._styles[col_name] = {}

                        # Normalizar filas
                        if isinstance(row_key, slice):
                            row_indices = list(range(row_key.start or 0,
                                                    row_key.stop or len(base_iloc.obj)))
                        elif isinstance(row_key, (list, np.ndarray)):
                            row_indices = list(row_key)
                        elif isinstance(row_key, int):
                            row_indices = [row_key]
                        else:
                            row_indices = []

                        # Guardar estilo por cada fila afectada
                        for idx in row_indices:
                            base_iloc.obj._styles[col_name][idx] = style

                        # Asignar datos al DataFrame
                        base_iloc[key] = data
                    else:
                        base_iloc[key] = data
                else:
                    # Caso normal: solo datos
                    base_iloc[key] = value

                # --- Sincronizar Excel ---
                try:
                    for j, col_name in enumerate(columns):
                        for i, val in enumerate(base_iloc.obj[col_name]):
                            cell = ws.cell(row=i+2, column=j+1, value=val)

                            # Aplicar estilo si existe
                            if (hasattr(base_iloc.obj, "_styles") and
                                col_name in base_iloc.obj._styles and
                                i in base_iloc.obj._styles[col_name]):
                                style = base_iloc.obj._styles[col_name][i]
                                base_iloc.obj._apply_style(cell, style)

                except Exception as e:
                    print(f"[ERROR] No se pudo actualizar Excel desde iloc: {e}")

            def __getattr__(_, name):
                return getattr(base_iloc, name)

        return _CustomILoc()



        # Sobrescribir __setitem__ (asignación directa de columnas)
    def __setitem__(self, key, value):
        if isinstance(value, dict) and "data" in value and "style" in value:
            style = value["style"]
            value = value["data"]

            # Guardar estilo en self._styles usando el nombre de la columna
            self._styles[key] = {"global":style}

        result = super().__setitem__(key, value)

        try:
            # Encontrar índice de columna en Excel
            if isinstance(key, str):
                col_excel = list(self.columns).index(key) + 1
            else:
                col_excel = key + 1

            # Volcar toda la columna actualizada
            for i, val in enumerate(self[key]):
                row_excel = i + 2  # +2 por cabecera
                self._ws.cell(row=row_excel, column=col_excel, value=val)

        except Exception as e:
            print(f"[ERROR] No se pudo actualizar Excel: {e}")

        return result

    # Sobrescribir _set_value
    def _set_value(self, index, col, value, takeable: bool = False):
        if isinstance(value, dict) and "data" in value and "style" in value:
            style = value["style"]
            value = value["data"]
            # Normalizar col: si es índice numérico, convertir a nombre de columna
            if isinstance(col, int):
                col_name = self.columns[col]
            else:
                col_name = col
            # Guardar estilo en self._styles usando el nombre de la columna
            self._styles[col_name][index] = style

        result = super()._set_value(index, col, value, takeable=takeable)

        try:
            # Volcar toda la fila actualizada
            row_excel = index + 2
            for j, col_name in enumerate(self.columns):
                val = self.at[index, col_name]
                self._ws.cell(row=row_excel, column=j+1, value=val)

        except Exception as e:
            print(f"[ERROR] No se pudo actualizar Excel desde _set_value: {e}")

        return result
    
    def sort_values(self, *args, **kwargs):
        ignore_index = kwargs.pop("ignore_index", False)

        old_index = self.index.copy()
        result = super().sort_values(*args, ignore_index=False, **kwargs)
        new_index = result.index

        # Construir mapeo nuevo -> viejo
        mapping = dict(zip(new_index, old_index))

        # Remapear estilos y asignarlos explícitamente
        remapped = self._remap_style_keys(mapping)
        object.__setattr__(result, "_styles", remapped)

        # Copiar otros metadatos
        for name in self._metadata:
            if name != "_styles":
                object.__setattr__(result, name, getattr(self, name, None))

        if ignore_index:
            result = result.reset_index(drop=True)

        return result

    def sort_index(self, *args, **kwargs):
        ignore_index = kwargs.pop("ignore_index", False)

        old_index = self.index.copy()
        result = super().sort_index(*args, ignore_index=False, **kwargs)
        new_index = result.index

        # Construir mapeo nuevo -> viejo
        mapping = dict(zip(new_index, old_index))

        # Remapear estilos
        remapped = self._remap_style_keys(mapping)
        object.__setattr__(result, "_styles", remapped)

        # Copiar otros metadatos
        for name in self._metadata:
            if name != "_styles":
                object.__setattr__(result, name, getattr(self, name, None))

        if ignore_index:
            result = result.reset_index(drop=True)

        return result

    def reindex(self, *args, **kwargs):
        old_index = self.index.copy()
        result = super().reindex(*args, **kwargs)
        new_index = result.index

        mapping = dict(zip(new_index, old_index))

        remapped = self._remap_style_keys(mapping)
        object.__setattr__(result, "_styles", remapped)

        for name in self._metadata:
            if name != "_styles":
                object.__setattr__(result, name, getattr(self, name, None))

        return result
        
    def sample(self, *args, **kwargs):
        old_index = self.index.copy()
        result = super().sample(*args, **kwargs)
        new_index = result.index

        mapping = dict(zip(new_index, old_index))

        remapped = self._remap_style_keys(mapping)
        object.__setattr__(result, "_styles", remapped)

        for name in self._metadata:
            if name != "_styles":
                object.__setattr__(result, name, getattr(self, name, None))

        return result

    def _remap_style_keys(self, old_to_new):
        new_styles = {}

        for col, rules in self._styles.items():
            temp_rules = {}

            # 1) Mover claves numéricas a su destino temporal
            for row_key, style in rules.items():
                if isinstance(row_key, int) and row_key in old_to_new:
                    dest = old_to_new[row_key]
                    temp_rules[f"{dest}_temp"] = style
                else:
                    temp_rules[row_key] = style

            # 2) Limpiar temporales
            final_rules = {}
            for row_key, style in temp_rules.items():
                if isinstance(row_key, str) and row_key.endswith("_temp"):
                    new_row = int(row_key[:-5])
                    final_rules[new_row] = style
                else:
                    final_rules[row_key] = style

            new_styles[col] = final_rules

        return new_styles

    def drop(self, labels=None, axis=0, index=None, columns=None, inplace=False, **kwargs):
        # Resolver qué se está borrando: filas (axis=0) o columnas (axis=1)
        is_cols = (axis == 1) or (columns is not None)
        is_rows = (axis == 0) or (index is not None)

        # Resolver conjuntos a eliminar (labels normalizados)
        cols_to_remove = None
        rows_to_remove = None

        if is_cols:
            cols_to_remove = columns if columns is not None else labels
            if isinstance(cols_to_remove, (str, int)):
                cols_to_remove = [cols_to_remove]
            elif cols_to_remove is None:
                cols_to_remove = []

        if is_rows:
            rows_to_remove = index if index is not None else labels
            if isinstance(rows_to_remove, (str, int)):
                rows_to_remove = [rows_to_remove]
            elif rows_to_remove is None:
                rows_to_remove = []

        # 1) CAPTURAR POSICIONES EN EL WORKSHEET ANTES DEL DROP
        col_positions = []
        row_positions = []

        if self._ws is not None:
            # Encabezados en fila 1 para columnas
            if cols_to_remove:
                header_values = [cell.value for cell in self._ws[1]]
                for name in cols_to_remove:
                    # Buscar por nombre exacto en encabezado
                    if name in header_values:
                        j = header_values.index(name) + 1  # Excel es 1-based
                        col_positions.append(j)
                    else:
                        # Fallback: usar la posición en el DF si existe
                        if name in list(self.columns):
                            j = list(self.columns).index(name) + 1
                            col_positions.append(j)

            # Filas de datos: mapeo por posición del índice actual
            if rows_to_remove:
                # Convertir labels del índice a posiciones enteras
                # Si el índice es RangeIndex, esto suele ser directo; si no, usamos get_indexer
                pos = self.index.get_indexer(rows_to_remove)
                # Filtrar -1 (no encontrados)
                pos = [p for p in pos if p != -1]
                # Convertir a posiciones en Excel (+2 por encabezado)
                row_positions = [p + 2 for p in pos]

        # 2) LIMPIAR EL WORKSHEET PRIMERO (EVITAR BÚSQUEDAS TRAS DROP)
        if self._ws is not None:
            # Borrar columnas en orden descendente para no desalinear índices
            for j in sorted(set(col_positions), reverse=True):
                self._ws.delete_cols(j)
            # Borrar filas en orden descendente
            for r in sorted(set(row_positions), reverse=True):
                self._ws.delete_rows(r)

        # 3) EJECUTAR EL DROP DE PANDAS
        result = super().drop(labels=labels, axis=axis, index=index, columns=columns, inplace=inplace, **kwargs)
        target = self if inplace else result

        # 4) LIMPIAR ESTILOS EN _styles
        if hasattr(target, "_styles"):
            # Columnas
            if cols_to_remove:
                for col_name in cols_to_remove:
                    target._styles.pop(col_name, None)

            # Filas (por índice de DF, no posiciones Excel)
            if rows_to_remove:
                for col_name, rules in list(target._styles.items()):
                    for r_label in rows_to_remove:
                        # Si guardas estilos por posición entera (0..n), necesitas convertir label->pos como arriba.
                        # Aquí asumimos que usas índices de fila (enteros) como claves.
                        rules.pop(r_label, None)

        return target



    def set_column_style(self, col_name, style: dict):
        """Aplica un estilo global a toda la columna."""
        if not hasattr(self, "_styles"):
            self._styles = {}
        if col_name not in self._styles:
            self._styles[col_name] = {}
        self._styles[col_name]["global"] = style
    
    def set_cell_style(self, row_idx: int, col_name: str, style: dict):
        """Aplica un estilo a una celda específica (fila, columna)."""
        if not hasattr(self, "_styles"):
            self._styles = {}
        if col_name not in self._styles:
            self._styles[col_name] = {}
        self._styles[col_name][row_idx] = style

    def set_range_style(self, row_slice: slice, col_name: str, style: dict):
        """Aplica un estilo a un rango de filas en una columna."""
        if not hasattr(self, "_styles"):
            self._styles = {}
        if col_name not in self._styles:
            self._styles[col_name] = {}
        self._styles[col_name][row_slice] = style

    def set_row_style(self, row_idx: int, style: dict):
        """Aplica un estilo a toda la fila (todas las columnas)."""
        if not hasattr(self, "_styles"):
            self._styles = {}
        for col_name in self.columns:
            if col_name not in self._styles:
                self._styles[col_name] = {}
            self._styles[col_name][row_idx] = style

    def set_header_row_style(self, style: dict):
        """Aplica un estilo a toda la fila de encabezados (fila 0 en Excel)."""
        if not hasattr(self, "_styles"):
            self._styles = {}
        for col_name in self.columns:
            if col_name not in self._styles:
                self._styles[col_name] = {}
            # Usamos un índice especial, por ejemplo "header"
            self._styles[col_name]["header"] = style

    def set_header_cell_style(self, col_name: str, style: dict):
        """Aplica un estilo a la celda de encabezado de una columna específica."""
        if not hasattr(self, "_styles"):
            self._styles = {}
        if col_name not in self._styles:
            self._styles[col_name] = {}
        self._styles[col_name]["header"] = style

    def set_global_style(self, style: dict):
        """Aplica un estilo global a todas las celdas del documento (encabezados y datos)."""
        if not hasattr(self, "_styles"):
            self._styles = {}
        # Usamos una clave especial "__document__"
        self._styles["__document__"] = {"global": style}


__all__ = ["DataFrameXL"]
